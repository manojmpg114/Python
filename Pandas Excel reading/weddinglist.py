import openpyxl

path = "/Users/manojgeorge/Documents/Python/Pandas Excel reading/seating_chart_data.xlsx"

wb = openpyxl.load_workbook(path)

sheet_obj = wb.active

cell_obj = sheet_obj.cell(row = 1, column = 2)

print(cell_obj.value)

total_rows = 77

print(total_rows)

name_list = []
for i in range(1, total_rows +1):
    name_cell = sheet_obj.cell(row = i,column = 1)
    table_cell = sheet_obj.cell(row = i, column = 2)
    #print(cell_obj.value)
    #name_list.append(name_cell.value + " " + table_cell)

    #print(name_cell.value + ": Table " + str(table_cell.value))

    #print(table_cell.value)
    #works
    #name_list.append(name_cell.value + ": Table " + str(table_cell.value))
    name_list.append("{:<19}".format(name_cell.value) + ": Table " + str(table_cell.value))

print(name_list)